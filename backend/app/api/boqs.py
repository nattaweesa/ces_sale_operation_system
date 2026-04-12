from __future__ import annotations
from datetime import datetime, timezone
from decimal import Decimal
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
import openpyxl

from app.database import get_db
from app.models.boq import BOQ, BOQItem
from app.models.deal import Deal
from app.models.product import Product
from app.models.user import User
from app.schemas.boq import (
    BOQCreate,
    BOQOut,
    BOQItemCreate,
    BOQItemUpdate,
    BOQItemOut,
    BOQImportPreviewOut,
    BOQImportPreviewRow,
)
from app.services.auth import get_current_user
from app.services.rbac import can_user

router = APIRouter(prefix="/boqs", tags=["boqs"])


async def _get_boq_or_404(boq_id: int, db: AsyncSession, current_user: User) -> BOQ:
    stmt = select(BOQ).where(BOQ.id == boq_id)
    if not await can_user(db, current_user, "boqs.view_all"):
        stmt = stmt.join(Deal, BOQ.deal_id == Deal.id).where(Deal.owner_id == current_user.id)

    result = await db.execute(stmt)
    boq = result.scalar_one_or_none()
    if not boq:
        raise HTTPException(status_code=404, detail="BOQ not found")
    return boq


def _cell_text(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _cell_num(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return float(v)
    txt = str(v).strip().replace(",", "")
    if not txt:
        return None
    try:
        return float(txt)
    except ValueError:
        return None


def _is_unit_like(v: object) -> bool:
    txt = _cell_text(v).lower().replace(" ", "")
    return txt in {"ea", "ea.", "pcs", "pc", "set", "lot", "point", "m", "meter", "unit"}


def _looks_item_code(v: object) -> bool:
    txt = _cell_text(v)
    if not txt:
        return False
    if len(txt) < 3:
        return False
    has_alpha = any(c.isalpha() for c in txt)
    has_digit = any(c.isdigit() for c in txt)
    return has_alpha and has_digit


def _is_section_heading(c0: object, c1: object, c2: object) -> bool:
    if _cell_num(c0) is not None:
        return False
    text = _cell_text(c1) or _cell_text(c0)
    if not text:
        return False
    # Template headings are often like "KNX-Merten Control Panel : B1 ..."
    lower = text.lower()
    return (
        "control panel" in lower
        or lower.startswith("central control")
        or lower.startswith("boq")
        or (":" in text and _cell_text(c2) == "")
    )


def _parse_boq_rows(rows: list[tuple[object, ...]]) -> tuple[str, list[dict], int, int]:
    # Auto-detect header row
    header_idx = None
    col_map: dict[str, int] = {}
    for ridx, r in enumerate(rows[:30]):
        header_row = [str(c).strip().lower() if c else "" for c in r]
        probe_map: dict[str, int] = {}
        for i, h in enumerate(header_row):
            if "seq" in h or "no" in h or "#" in h:
                probe_map.setdefault("seq", i)
            elif "desc" in h:
                probe_map["description"] = i
            elif "qty" in h or "quantity" in h:
                probe_map["quantity"] = i
            elif "unit" in h:
                probe_map["unit"] = i
            elif "section" in h or "area" in h or "zone" in h:
                probe_map["section_label"] = i
        if "description" in probe_map:
            header_idx = ridx
            col_map = probe_map
            break

    parsed_rows: list[dict] = []
    skipped_rows = 0

    if header_idx is not None:
        mode = "header"
        seq = 1
        for row in rows[header_idx + 1:]:
            desc = row[col_map["description"]] if "description" in col_map else None
            if not _cell_text(desc):
                skipped_rows += 1
                continue

            qty = 1
            if "quantity" in col_map:
                qty = _cell_num(row[col_map["quantity"]]) or 1

            unit = None
            if "unit" in col_map and _cell_text(row[col_map["unit"]]):
                unit = _cell_text(row[col_map["unit"]])

            section_label = None
            if "section_label" in col_map and _cell_text(row[col_map["section_label"]]):
                section_label = _cell_text(row[col_map["section_label"]])

            parsed_rows.append(
                {
                    "seq": int(_cell_num(row[col_map["seq"]]) or seq) if "seq" in col_map else seq,
                    "description": _cell_text(desc),
                    "quantity": qty,
                    "unit": unit,
                    "section_label": section_label,
                    "notes": None,
                }
            )
            seq += 1
    else:
        mode = "template"
        current_section: Optional[str] = None

        for row in rows:
            c0 = row[0] if len(row) > 0 else None
            c1 = row[1] if len(row) > 1 else None
            c2 = row[2] if len(row) > 2 else None
            c3 = row[3] if len(row) > 3 else None
            c4 = row[4] if len(row) > 4 else None

            if _is_section_heading(c0, c1, c2):
                sec = _cell_text(c1) or _cell_text(c0)
                if sec and not sec.lower().startswith("remark"):
                    current_section = sec
                continue

            seq_num = _cell_num(c0)
            if seq_num is None:
                skipped_rows += 1
                continue

            text_c1 = _cell_text(c1)
            text_c2 = _cell_text(c2)

            description = text_c2 if text_c2 else text_c1
            if not description:
                skipped_rows += 1
                continue

            qty = _cell_num(c4)
            if qty is None:
                qty = _cell_num(c2)
            if qty is None:
                qty = 1

            unit = _cell_text(c3) if _is_unit_like(c3) else None

            notes = None
            if text_c2 and _looks_item_code(c1):
                notes = f"Item code: {text_c1}"

            parsed_rows.append(
                {
                    "seq": int(seq_num),
                    "description": description,
                    "quantity": qty,
                    "unit": unit,
                    "section_label": current_section,
                    "notes": notes,
                }
            )

    total_rows = max(len(rows) - (header_idx + 1 if header_idx is not None else 0), 0)
    return mode, parsed_rows, skipped_rows, total_rows


def _enrich_item(item: BOQItem) -> BOQItemOut:
    out = BOQItemOut.model_validate(item)
    if item.product:
        out.product_code = item.product.item_code
        out.product_description = item.product.description
    return out


@router.get("/import-template/download")
async def download_boq_template(_: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ Template"

    ws.append(["Project : <Project Name>"])
    ws.append(["Remark : You can keep this format and add your own sections/items"])
    ws.append([])
    ws.append(["Central control"])
    ws.append([1, "LSS100200", "SPACELYNK LOGIC CONTROLLER", "ea.", 1])
    ws.append([2, "MTN6513-1201", "KNX power supply", "ea.", 2])
    ws.append([])
    ws.append(["KNX-Merten Control Panel : B1"])
    ws.append([1, "MTN6705-0008", "SpaceLogic KNX Uni. Switch Master 8ch", "Merten", 1, 1000, 1000])
    ws.append([2, "MTN6805-0008", "SpaceLogic KNX Uni. Switch Extension 8ch", "Merten", 2, 1000, 2000])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=BOQ_Import_Template.xlsx"},
    )


@router.post("/import-preview", response_model=BOQImportPreviewOut)
async def preview_boq_import(
    file: UploadFile = File(...),
    _: User = Depends(get_current_user),
):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty spreadsheet")

    mode, parsed_rows, skipped_rows, total_rows = _parse_boq_rows(rows)

    sample_rows = [BOQImportPreviewRow(**r) for r in parsed_rows[:20]]
    return BOQImportPreviewOut(
        mode=mode,
        total_rows=total_rows,
        imported_rows=len(parsed_rows),
        skipped_rows=skipped_rows,
        sample_rows=sample_rows,
    )


@router.post("", response_model=BOQOut, status_code=201)
async def create_boq(payload: BOQCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if payload.deal_id:
        deal_res = await db.execute(select(Deal).where(Deal.id == payload.deal_id))
        deal = deal_res.scalar_one_or_none()
        if not deal:
            raise HTTPException(status_code=404, detail="Deal not found")
        if not await can_user(db, current_user, "boqs.view_all") and deal.owner_id != current_user.id:
            raise HTTPException(status_code=403, detail="You can only create BOQ for your own deals")
        if not deal.project_id:
            raise HTTPException(status_code=400, detail="Selected deal has no project")
        if deal.project_id != payload.project_id:
            raise HTTPException(status_code=400, detail="Deal and project mismatch")

        existing_res = await db.execute(select(BOQ).where(BOQ.deal_id == payload.deal_id))
        existing = existing_res.scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=409, detail=f"This deal already has BOQ id={existing.id}")

    boq = BOQ(project_id=payload.project_id, deal_id=payload.deal_id, name=payload.name)
    db.add(boq)
    await db.commit()
    stmt = select(BOQ).options(selectinload(BOQ.items).selectinload(BOQItem.product)).where(BOQ.id == boq.id)
    result = await db.execute(stmt)
    return result.scalar_one()


@router.get("", response_model=list[BOQOut])
async def list_boqs(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    stmt = (
        select(BOQ)
        .options(selectinload(BOQ.items).selectinload(BOQItem.product))
        .order_by(BOQ.created_at.desc())
    )
    if not await can_user(db, current_user, "boqs.view_all"):
        stmt = stmt.join(Deal, BOQ.deal_id == Deal.id).where(Deal.owner_id == current_user.id)
    result = await db.execute(stmt)
    return list(result.scalars().all())


@router.get("/{boq_id}", response_model=BOQOut)
async def get_boq(boq_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await _get_boq_or_404(boq_id, db, current_user)
    stmt = select(BOQ).options(selectinload(BOQ.items).selectinload(BOQItem.product)).where(BOQ.id == boq_id)
    result = await db.execute(stmt)
    return result.scalar_one()


@router.post("/{boq_id}/items", response_model=BOQItemOut, status_code=201)
async def add_boq_item(boq_id: int, payload: BOQItemCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await _get_boq_or_404(boq_id, db, current_user)

    item = BOQItem(boq_id=boq_id, **payload.model_dump())
    if payload.product_id:
        item.mapped_at = datetime.now(timezone.utc)
    db.add(item)
    await db.commit()
    stmt = select(BOQItem).options(selectinload(BOQItem.product)).where(BOQItem.id == item.id)
    result2 = await db.execute(stmt)
    return _enrich_item(result2.scalar_one())


@router.put("/{boq_id}/items/{item_id}", response_model=BOQItemOut)
async def update_boq_item(boq_id: int, item_id: int, payload: BOQItemUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await _get_boq_or_404(boq_id, db, current_user)
    stmt = select(BOQItem).options(selectinload(BOQItem.product)).where(BOQItem.id == item_id, BOQItem.boq_id == boq_id)
    result = await db.execute(stmt)
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="BOQ item not found")

    prev_product_id = item.product_id
    for field, val in payload.model_dump(exclude_none=True).items():
        setattr(item, field, val)
    if payload.product_id and payload.product_id != prev_product_id:
        item.mapped_at = datetime.now(timezone.utc)
    await db.commit()
    stmt2 = select(BOQItem).options(selectinload(BOQItem.product)).where(BOQItem.id == item_id)
    result2 = await db.execute(stmt2)
    return _enrich_item(result2.scalar_one())


@router.delete("/{boq_id}/items/{item_id}", status_code=204)
async def delete_boq_item(boq_id: int, item_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await _get_boq_or_404(boq_id, db, current_user)
    result = await db.execute(select(BOQItem).where(BOQItem.id == item_id, BOQItem.boq_id == boq_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="BOQ item not found")
    await db.delete(item)
    await db.commit()


@router.post("/{boq_id}/import", response_model=BOQOut)
async def import_boq_from_excel(
    boq_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import BOQ items from Excel. Expected columns: seq, description, quantity, unit, section_label"""
    boq = await _get_boq_or_404(boq_id, db, current_user)

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty spreadsheet")
    _, parsed_rows, _, _ = _parse_boq_rows(rows)
    imported = 0
    for row in parsed_rows:
        item = BOQItem(boq_id=boq_id, **row)
        db.add(item)
        imported += 1

    if imported == 0:
        raise HTTPException(
            status_code=400,
            detail="No BOQ item rows found. Please use BOQ template format (with sequence + description).",
        )

    boq.source = "excel_import"
    await db.commit()

    stmt = select(BOQ).options(selectinload(BOQ.items).selectinload(BOQItem.product)).where(BOQ.id == boq_id)
    r = await db.execute(stmt)
    return r.scalar_one()
