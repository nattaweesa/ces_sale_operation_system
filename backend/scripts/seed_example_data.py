"""
Seed script: imports all ExampleData files into the CES system.

Creates:
  - Brands: Merten, Schneider, Delta
  - Categories: KNX Lighting Control, C-Bus Lighting Control, RCU, Accessories
  - Customers + Contacts (5 customers from project names)
  - Projects (5 projects linked to customers)
  - BOQs + BOQ items from all 7 Excel files
  - Quotations with correct totals from 3 PDF files
  - Copies PDFs to storage/imports as project attachments (as audit trail)

Run from backend/ folder:
    PYTHONPATH=. python scripts/seed_example_data.py
"""
from __future__ import annotations

import asyncio
import os
import re
import shutil
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy import select

from app.database import AsyncSessionLocal
from app.models.brand import Brand
from app.models.category import Category
from app.models.customer import Customer, Contact
from app.models.project import Project
from app.models.boq import BOQ, BOQItem
from app.models.quotation import (
    Quotation, QuotationSection, QuotationLine, QuotationRevision,
)
from app.models.user import User

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR.parent / "ExampleData"
STORAGE_DIR = BASE_DIR / "storage"


# ──────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────

async def upsert_brand(session, name: str) -> Brand:
    q = await session.execute(select(Brand).where(Brand.name == name))
    obj = q.scalar_one_or_none()
    if not obj:
        obj = Brand(name=name)
        session.add(obj)
        await session.flush()
    return obj


async def upsert_category(session, name: str, parent_id: int | None = None) -> Category:
    q = await session.execute(select(Category).where(Category.name == name))
    obj = q.scalar_one_or_none()
    if not obj:
        obj = Category(name=name, parent_id=parent_id)
        session.add(obj)
        await session.flush()
    return obj


async def upsert_customer(session, name: str, **kwargs) -> Customer:
    q = await session.execute(select(Customer).where(Customer.name == name))
    obj = q.scalar_one_or_none()
    if not obj:
        obj = Customer(name=name, **kwargs)
        session.add(obj)
        await session.flush()
    return obj


async def create_contact(session, customer_id: int, full_name: str, title: str, phone: str | None, email: str | None, is_primary: bool) -> Contact:
    q = await session.execute(
        select(Contact).where(Contact.customer_id == customer_id, Contact.full_name == full_name)
    )
    obj = q.scalar_one_or_none()
    if not obj:
        obj = Contact(customer_id=customer_id, full_name=full_name, title=title,
                      phone=phone, email=email, is_primary=is_primary)
        session.add(obj)
        await session.flush()
    return obj


async def upsert_project(session, name: str, customer_id: int, **kwargs) -> Project:
    q = await session.execute(select(Project).where(Project.name == name))
    obj = q.scalar_one_or_none()
    if not obj:
        obj = Project(name=name, customer_id=customer_id, **kwargs)
        session.add(obj)
        await session.flush()
    return obj


async def create_boq(session, project_id: int, name: str, source: str = "excel_import") -> BOQ:
    q = await session.execute(select(BOQ).where(BOQ.project_id == project_id, BOQ.name == name))
    obj = q.scalar_one_or_none()
    if not obj:
        obj = BOQ(project_id=project_id, name=name, source=source)
        session.add(obj)
        await session.flush()
    return obj


def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val).replace(",", "").strip())
    except Exception:
        return None


def _clean(val) -> str:
    if val is None:
        return ""
    return " ".join(str(val).split()).strip()


# ──────────────────────────────────────────────────────
# BOQ Excel parsers
# ──────────────────────────────────────────────────────

def parse_knx_cbus_boq(filepath: Path) -> list[dict]:
    """
    Parse KNX or C-Bus BOQ Excel.
    Expected cols: 0=item#, 1=cat_no, 2=description, 3=brand, 4=qty, 5=unit_price, 6=total
    Section headers are rows where col0 has text but col1/col2 are empty, or col0 is None
    and there is text in col2 but not parseable as item.
    """
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    items = []
    current_section = None
    seq = 0

    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue

        c0 = row[0] if len(row) > 0 else None
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None
        c3 = row[3] if len(row) > 3 else None
        c4 = row[4] if len(row) > 4 else None
        c5 = row[5] if len(row) > 5 else None

        # Detect section heading: non-numeric c0 with no c1, OR c0 is None but c2 exists and looks like heading
        item_num = None
        if c0 is not None:
            try:
                item_num = int(float(str(c0).strip()))
            except (ValueError, TypeError):
                pass

        if item_num is None:
            # Check if this is a section heading
            heading_text = _clean(c0) or _clean(c1) or _clean(c2)
            if heading_text and not heading_text.lower().startswith(("remark", "หมาย", "total", "Price list", "www", "complete", "bangkok", "phuket", "chiang", "1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.")):
                current_section = heading_text[:150]
            continue

        desc = _clean(c2) or _clean(c1)
        if not desc:
            continue

        qty = _to_decimal(c4) or Decimal("1")
        unit_price = _to_decimal(c5) or Decimal("0")
        brand = _clean(c3) or "Unknown"

        # Skip rows with qty=0 and price=0 (placeholders)
        if qty == 0 and unit_price == 0:
            continue

        seq += 1
        items.append({
            "seq": seq,
            "description": desc,
            "item_code": _clean(c1) or None,
            "quantity": qty,
            "unit": "ea.",
            "unit_price": unit_price,
            "section_label": current_section,
            "brand": brand,
            "notes": None,
        })

    return items


def parse_rcu_boq(filepath: Path, sheet_idx: int = 0) -> list[dict]:
    """
    Parse RCU BOQ Excel.
    Expected cols: 0=item#, 1=description, 2=brand, 3=model(item_code), 4=qty
    """
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.worksheets[sheet_idx]
    items = []
    current_section = None
    seq = 0

    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue

        c0 = row[0] if len(row) > 0 else None
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None
        c3 = row[3] if len(row) > 3 else None
        c4 = row[4] if len(row) > 4 else None
        c5 = row[5] if len(row) > 5 else None

        item_num = None
        if c0 is not None:
            try:
                item_num = int(float(str(c0).strip()))
            except (ValueError, TypeError):
                pass

        if item_num is None:
            heading = _clean(c0) or _clean(c1)
            if heading and not heading.lower().startswith(("remark", "หมาย", "total", "www", "complete", "bangkok", "ref", "date", "project", "1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "10.")):
                current_section = heading[:150]
            continue

        desc = _clean(c1)
        if not desc:
            continue

        # For RCU files, qty can be in col4, or in c3 if model is missing
        qty = _to_decimal(c4) or Decimal("1")
        brand = _clean(c2) or "SMART"
        item_code = _clean(c3) or None
        notes = _clean(c5) if c5 else None

        if qty == 0:
            qty = Decimal("1")

        seq += 1
        items.append({
            "seq": seq,
            "description": desc,
            "item_code": item_code,
            "quantity": qty,
            "unit": "ea.",
            "unit_price": Decimal("0"),
            "section_label": current_section,
            "brand": brand,
            "notes": notes,
        })

    return items


# ──────────────────────────────────────────────────────
# Quotation builder (from PDF data — amounts hardcoded)
# ──────────────────────────────────────────────────────

async def get_next_qt_number(session) -> str:
    from sqlalchemy import func
    year = 2026
    prefix = "CES-QT"
    count_result = await session.execute(
        select(func.count(Quotation.id)).where(
            Quotation.quotation_number.like(f"{prefix}-{year}-%")
        )
    )
    count = count_result.scalar_one()
    return f"{prefix}-{year}-{count + 1:04d}"


async def create_quotation_record(
    session,
    project_id: int,
    boq_id: int | None,
    contact_id: int | None,
    sales_owner_id: int | None,
    subject: str,
    original_qt_number: str,
    subtotal: Decimal,
    vat_rate: Decimal,
    vat_amount: Decimal,
    grand_total: Decimal,
    lines: list[dict],
    sections: list[str],
    delivery_terms: str,
    validity_days: int,
    payment_terms: str,
    remarks: list[str],
) -> Quotation:
    qt_number = await get_next_qt_number(session)

    qt = Quotation(
        quotation_number=qt_number,
        project_id=project_id,
        contact_id=contact_id,
        sales_owner_id=sales_owner_id,
        subject=subject,
        status="issued",
        current_revision=1,
        subtotal=subtotal,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        grand_total=grand_total,
        delivery_terms=delivery_terms,
        validity_days=validity_days,
        payment_terms=payment_terms,
        internal_notes=f"Original Ref: {original_qt_number}",
        scope_of_work="\n".join(remarks) if remarks else None,
        exclusions=None,
        boq_id=boq_id,
    )
    session.add(qt)
    await session.flush()

    # Create sections
    section_objs = {}
    for i, sec_label in enumerate(sections):
        sec = QuotationSection(quotation_id=qt.id, label=sec_label, sort_order=i)
        session.add(sec)
        await session.flush()
        section_objs[sec_label] = sec.id

    # Create lines
    for i, line_data in enumerate(lines):
        sec_label = line_data.get("section_label")
        net_price = line_data.get("unit_price", Decimal("0"))
        qty = line_data.get("quantity", Decimal("1"))
        amount = net_price * qty

        line = QuotationLine(
            quotation_id=qt.id,
            section_id=section_objs.get(sec_label) if sec_label else None,
            seq=i + 1,
            product_id=None,
            item_code=line_data.get("item_code"),
            description=line_data.get("description", ""),
            brand=line_data.get("brand"),
            list_price=net_price,
            discount_pct=Decimal("0"),
            net_price=net_price,
            quantity=qty,
            unit=line_data.get("unit", "ea."),
            amount=amount,
            remark=line_data.get("notes"),
            is_optional=False,
        )
        session.add(line)

    # Create revision snapshot
    snapshot = {
        "quotation_number": qt_number,
        "original_ref": original_qt_number,
        "subject": subject,
        "subtotal": str(subtotal),
        "vat_rate": str(vat_rate),
        "vat_amount": str(vat_amount),
        "grand_total": str(grand_total),
        "lines_count": len(lines),
    }
    rev = QuotationRevision(
        quotation_id=qt.id,
        revision_number=1,
        issued_by=sales_owner_id,
        snapshot_json=snapshot,
        pdf_path=None,
    )
    session.add(rev)
    await session.flush()
    return qt


# ──────────────────────────────────────────────────────
# Main seed
# ──────────────────────────────────────────────────────

async def main():
    async with AsyncSessionLocal() as session:

        print("\n── Brands ──")
        brand_smart = await upsert_brand(session, "SMART")
        brand_merten = await upsert_brand(session, "Merten")
        brand_schneider = await upsert_brand(session, "Schneider")
        brand_delta = await upsert_brand(session, "Delta")
        print(f"  SMART id={brand_smart.id}, Merten id={brand_merten.id}, Schneider id={brand_schneider.id}, Delta id={brand_delta.id}")

        print("\n── Categories ──")
        cat_knx = await upsert_category(session, "KNX Lighting Control")
        cat_cbus = await upsert_category(session, "C-Bus Lighting Control")
        cat_rcu = await upsert_category(session, "RCU (Room Control Unit)")
        cat_acc = await upsert_category(session, "Wiring Accessories")
        print(f"  KNX id={cat_knx.id}, C-Bus id={cat_cbus.id}, RCU id={cat_rcu.id}")

        print("\n── Customers ──")
        cust_sangfa = await upsert_customer(session, "แสงฟ้าก่อสร้าง",
            address="Bangkok", industry="Construction", is_active=True)
        cust_stepwise = await upsert_customer(session, "Stepwise",
            address="Bangkok", industry="Interior Design", is_active=True)
        cust_ylg = await upsert_customer(session, "YLG",
            address="Bangkok", industry="Commercial", is_active=True)
        cust_mx27 = await upsert_customer(session, "MX 27 Hotel",
            address="Bangkok", industry="Hospitality", is_active=True)
        cust_angel = await upsert_customer(session, "Angel Hotel (Pakchong)",
            address="Pakchong, Nakhon Ratchasima", industry="Hospitality", is_active=True)
        print(f"  5 customers created/found")

        print("\n── Contacts ──")
        contact_atthaphon = await create_contact(session, cust_sangfa.id,
            "คุณอรรถพล", "คุณ", None, None, is_primary=True)
        contact_rattanapon = await create_contact(session, cust_stepwise.id,
            "คุณรัตนพล", "คุณ", None, None, is_primary=True)
        print(f"  2 contacts created/found")

        print("\n── Sales user ──")
        q = await session.execute(select(User).where(User.email == "phiromrat@ces-asia.com"))
        sales_user = q.scalar_one_or_none()
        if not sales_user:
            q = await session.execute(select(User).where(User.username == "phiromrat"))
            sales_user = q.scalar_one_or_none()
        if not sales_user:
            q = await session.execute(select(User).limit(1))
            sales_user = q.scalar_one_or_none()
        print(f"  Using sales user id={sales_user.id if sales_user else None}")

        print("\n── Projects ──")
        proj_hyatt = await upsert_project(session, "Hyatt House Asoke", cust_sangfa.id,
            location="Asoke, Bangkok", description="KNX Merten Lighting Control System, Option 1", status="active")
        proj_sathon21 = await upsert_project(session, "Sathon 21 Project Bangkok", cust_stepwise.id,
            location="Sathorn, Bangkok", description="KNX Merten Lighting Control: Romsai Restaurant + Lobby Lounge", status="active")
        proj_ylg = await upsert_project(session, "YLG Office Building", cust_ylg.id,
            location="Bangkok", description="KNX Option 1 + C-Bus Option 2 (Lighting Control Systems)", status="active")
        proj_mx27 = await upsert_project(session, "MX 27 HOTEL", cust_mx27.id,
            location="Bangkok", description="RCU S-3000P System + IOT Integrator, 171 rooms", status="active")
        proj_angel = await upsert_project(session, "โรงแรมปากช่อง (Angel Hotel)", cust_angel.id,
            location="Pakchong, Nakhon Ratchasima", description="RCU S-1002B Budget Hotel System", status="active")
        print(f"  5 projects created/found")

        await session.commit()
        print("\n── BOQs ──")

        # ── BOQ 1: Hyatt House KNX ──
        boq_hyatt = await create_boq(session, proj_hyatt.id, "BOQ KNX Merten – Hyatt House Asoke Rev05 Option1")
        items_hyatt = parse_knx_cbus_boq(DATA_DIR / "2026-02-23_BOQ_KNX@โครงการ Hyatt House Asoke_Rev05-Option1.xlsx")
        for it in items_hyatt:
            session.add(BOQItem(
                boq_id=boq_hyatt.id, seq=it["seq"], description=it["description"],
                quantity=it["quantity"], unit=it["unit"], section_label=it["section_label"],
                notes=it["notes"],
            ))
        await session.flush()
        print(f"  BOQ Hyatt House: {len(items_hyatt)} items")

        # ── BOQ 2: Sathon 21 – Romsai Restaurant ──
        boq_romsai = await create_boq(session, proj_sathon21.id, "BOQ KNX – Sathon 21 Romsai Restaurant")
        items_romsai = parse_knx_cbus_boq(DATA_DIR / "3'19_2026_BOQ_KNX@Sathon 21 (ROMSAI RESTAURANT).xlsx")
        for it in items_romsai:
            session.add(BOQItem(
                boq_id=boq_romsai.id, seq=it["seq"], description=it["description"],
                quantity=it["quantity"], unit=it["unit"], section_label=it["section_label"],
                notes=it["notes"],
            ))
        await session.flush()
        print(f"  BOQ Sathon21 Romsai: {len(items_romsai)} items")

        # ── BOQ 3: Sathon 21 – Lobby Lounge ──
        boq_lobby = await create_boq(session, proj_sathon21.id, "BOQ KNX – Sathon 21 Lobby Lounge Rev3")
        items_lobby = parse_knx_cbus_boq(DATA_DIR / "REV.3_3'24_2026_BOQ_KNX@Sathon 21 (Lobby Lounge).xlsx")
        for it in items_lobby:
            session.add(BOQItem(
                boq_id=boq_lobby.id, seq=it["seq"], description=it["description"],
                quantity=it["quantity"], unit=it["unit"], section_label=it["section_label"],
                notes=it["notes"],
            ))
        await session.flush()
        print(f"  BOQ Sathon21 Lobby: {len(items_lobby)} items")

        # ── BOQ 4: YLG KNX Option 1 ──
        boq_ylg_knx = await create_boq(session, proj_ylg.id, "BOQ KNX Merten – YLG Office Option 1")
        items_ylg_knx = parse_knx_cbus_boq(DATA_DIR / "2026-02-16_BOQ_KNX@YLG_Option1.xlsx")
        for it in items_ylg_knx:
            session.add(BOQItem(
                boq_id=boq_ylg_knx.id, seq=it["seq"], description=it["description"],
                quantity=it["quantity"], unit=it["unit"], section_label=it["section_label"],
                notes=it["notes"],
            ))
        await session.flush()
        print(f"  BOQ YLG KNX Option1: {len(items_ylg_knx)} items")

        # ── BOQ 5: YLG C-Bus Option 2 ──
        boq_ylg_cbus = await create_boq(session, proj_ylg.id, "BOQ C-Bus – YLG Office Option 2")
        items_ylg_cbus = parse_knx_cbus_boq(DATA_DIR / "2026-02-16_BOQ_C-Bus @YLG_Option2.xlsx")
        for it in items_ylg_cbus:
            session.add(BOQItem(
                boq_id=boq_ylg_cbus.id, seq=it["seq"], description=it["description"],
                quantity=it["quantity"], unit=it["unit"], section_label=it["section_label"],
                notes=it["notes"],
            ))
        await session.flush()
        print(f"  BOQ YLG C-Bus Option2: {len(items_ylg_cbus)} items")

        # ── BOQ 6: MX 27 Hotel RCU ──
        boq_mx27 = await create_boq(session, proj_mx27.id, "BOQ RCU S-3000P – MX 27 HOTEL")
        items_mx27 = parse_rcu_boq(DATA_DIR / "26-3-25BOQ_RCU@MX 27 HOTEL.xlsx", sheet_idx=1)
        for it in items_mx27:
            session.add(BOQItem(
                boq_id=boq_mx27.id, seq=it["seq"], description=it["description"],
                quantity=it["quantity"], unit=it["unit"], section_label=it["section_label"],
                notes=it["notes"],
            ))
        await session.flush()
        print(f"  BOQ MX 27 Hotel: {len(items_mx27)} items")

        # ── BOQ 7: Angel Hotel RCU ──
        boq_angel = await create_boq(session, proj_angel.id, "BOQ RCU S-1002B – Angel Hotel (Pakchong)")
        items_angel = parse_rcu_boq(DATA_DIR / "26-3-25_BOQ RCU@5 STOREY  BUDGET HOTEL.xlsx", sheet_idx=0)
        for it in items_angel:
            session.add(BOQItem(
                boq_id=boq_angel.id, seq=it["seq"], description=it["description"],
                quantity=it["quantity"], unit=it["unit"], section_label=it["section_label"],
                notes=it["notes"],
            ))
        await session.flush()
        print(f"  BOQ Angel Hotel: {len(items_angel)} items")

        await session.commit()

        print("\n── Quotations ──")
        sales_id = sales_user.id if sales_user else None

        # ── QT 1: Hyatt House – Q_KNX_025R1-26PA ──
        remarks_hyatt = [
            "ตู้คอนโทรลจะต้องเชื่อมต่อถึงกันโดยสาย KNX เท่านั้น",
            "1. Control แบบ On/Off จำนวนโหลดต้องไม่เกิน 16A ต่อ 1 วงจร",
            "2. วงจรที่เป็น DIM PHASE ระบบ C-Bus จะคอนโทรลด้วย KNX",
            "3. ราคานี้นำเสนอตามเอกสารแบบ 31 OCT 2024 เท่านั้น",
            "4. ราคานี้รวม Computer Central control",
            "5. ราคานี้ไม่รวมคอนโทรลม่าน, แอร์ และ motion sensor",
            "6. ราคาที่นำเสนอนี้ไม่ได้เชื่อมต่อกับระบบ lighting control เดิม",
            "7. ราคาที่นำเสนอนี้ไม่รวมคอมพิวเตอร์ระบบ BAS",
            "8. ราคาที่นำเสนอนี้ไม่รวมงานติดตั้งและการเดินสาย Wiring",
            "9. ราคาที่นำเสนอนี้ไม่รวมตู้ B3.1, B3.2, B3.3, B3.4",
        ]
        qt_hyatt_lines = [
            {"seq": 1, "description": "KNX Merten Lighting Control System – Hyatt House Asoke Option 1 (as per BOQ Rev05)",
             "item_code": None, "brand": "Merten", "quantity": Decimal("1"),
             "unit": "Lot", "unit_price": Decimal("2676268.08"),
             "section_label": "KNX System", "notes": None},
        ]
        qt_hyatt = await create_quotation_record(
            session,
            project_id=proj_hyatt.id,
            boq_id=boq_hyatt.id,
            contact_id=contact_atthaphon.id,
            sales_owner_id=sales_id,
            subject="Quotation of KNX Merten System – Hyatt House Asoke Option 1",
            original_qt_number="Q_KNX_025R1-26PA",
            subtotal=Decimal("2676268.08"),
            vat_rate=Decimal("7.00"),
            vat_amount=Decimal("187338.77"),
            grand_total=Decimal("2863606.85"),
            lines=qt_hyatt_lines,
            sections=["KNX System"],
            delivery_terms="30-60 days after receipt of purchase order and down payment. Delivery to Bangkok area.",
            validity_days=30,
            payment_terms="40% Down payment by cashier cheque against purchase order. Balance 60% upon delivery.",
            remarks=remarks_hyatt,
        )
        print(f"  QT Hyatt House: {qt_hyatt.quotation_number} (Grand Total: {qt_hyatt.grand_total:,.2f})")

        # ── QT 2: Sathon 21 – Q_KNX_037R2-26PA (Romsai + extra) ──
        remarks_sathon_037 = [
            "1. Control แบบ On/Off, Phase dimming และ 1-10V dimming",
            "2. ราคาที่นำเสนอจะมีในส่วน Interior (ROMSAI Restaurant) เท่านั้น",
            "3. ราคาที่นำเสนอนี้อุปกรณ์ทุกตัวจะต้องเชื่อมต่อถึงกัน",
            "4. ราคานี้นำเสนอตามเอกสารของ Bo Steiber (ROMSAI Restaurant)",
            "5. ราคาที่นำเสนอนี้ไม่ได้เชื่อมต่อกับระบบ lighting control เดิม",
            "6. จำนวนสวิทช์อ้างอิงตามจำนวนของเอกสาร Bo Steiber",
            "7. ราคาที่นำเสนอนี้ไม่รวมงานติดตั้งและการเดินสาย Wiring",
        ]
        qt_037_lines = [
            {"seq": 1, "description": "KNX Merten Lighting Control System – Romsai Restaurant & Lift Lobby",
             "item_code": None, "brand": "Merten", "quantity": Decimal("1"),
             "unit": "Lot", "unit_price": Decimal("316986.00"),
             "section_label": "KNX System", "notes": "ROMSAI RESTAURANT & LIFT LOBBY + ROMSAI WC"},
            {"seq": 2, "description": "Switch Panels & Accessories – Romsai Restaurant",
             "item_code": None, "brand": "Merten", "quantity": Decimal("1"),
             "unit": "Lot", "unit_price": Decimal("29100.00"),
             "section_label": "Switch Panels", "notes": "Cuadro H push buttons"},
        ]
        qt_037 = await create_quotation_record(
            session,
            project_id=proj_sathon21.id,
            boq_id=boq_romsai.id,
            contact_id=contact_rattanapon.id,
            sales_owner_id=sales_id,
            subject="Quotation KNX Merten Lighting Control – Sathon 21 Romsai Restaurant (Rev 2)",
            original_qt_number="Q_KNX_037R2-26PA",
            subtotal=Decimal("346086.00"),
            vat_rate=Decimal("7.00"),
            vat_amount=Decimal("24226.02"),
            grand_total=Decimal("370312.02"),
            lines=qt_037_lines,
            sections=["KNX System", "Switch Panels"],
            delivery_terms="60-120 days after received purchase order and deposit. Delivery to Bangkok Area.",
            validity_days=15,
            payment_terms="Credit 60 Days.",
            remarks=remarks_sathon_037,
        )
        print(f"  QT Sathon21 R037: {qt_037.quotation_number} (Grand Total: {qt_037.grand_total:,.2f})")

        # ── QT 3: Sathon 21 – Q_KNX_039-26PA (Lobby Lounge) ──
        remarks_sathon_039 = [
            "1. Control 1-10V dimming จำนวนโหลดต้องไม่เกิน 10A ต่อ 1 วงจร",
            "2. ราคาที่นำเสนอจะมีในส่วน Interior Lobby Lounge เท่านั้น",
            "3. ราคาที่นำเสนอนี้อุปกรณ์ทุกตัวจะต้องเชื่อมต่อถึงกัน",
            "4. ราคานี้นำเสนอตามเอกสารของ Bo Steiber พื้นที่ Lobby Lounge",
            "5. ในส่วนของ Lobby Lounge โคมไฟ Decorative ไม่มีการ Dimming",
            "6. ราคาที่นำเสนอนี้ไม่ได้เชื่อมต่อกับระบบ lighting control เดิม",
            "7. จำนวนสวิทช์อ้างอิงตาม Stepwise",
            "8. ราคาที่นำเสนอนี้ไม่รวมงานติดตั้งและการเดินสาย Wiring",
        ]
        qt_039_lines = [
            {"seq": 1, "description": "KNX Merten Lighting Control System – Lobby Lounge (1-10V Dimming)",
             "item_code": None, "brand": "Merten", "quantity": Decimal("1"),
             "unit": "Lot", "unit_price": Decimal("286486.00"),
             "section_label": "KNX System", "notes": "L1 Lobby Lounge (2 panels)"},
        ]
        qt_039 = await create_quotation_record(
            session,
            project_id=proj_sathon21.id,
            boq_id=boq_lobby.id,
            contact_id=contact_rattanapon.id,
            sales_owner_id=sales_id,
            subject="Quotation KNX Merten Lighting Control – Sathon 21 Lobby Lounge",
            original_qt_number="Q_KNX_039-26PA",
            subtotal=Decimal("286486.00"),
            vat_rate=Decimal("7.00"),
            vat_amount=Decimal("20054.02"),
            grand_total=Decimal("306540.02"),
            lines=qt_039_lines,
            sections=["KNX System"],
            delivery_terms="60-120 days after received purchase order and deposit. Delivery to Bangkok Area.",
            validity_days=15,
            payment_terms="Credit 60 Days.",
            remarks=remarks_sathon_039,
        )
        print(f"  QT Sathon21 R039: {qt_039.quotation_number} (Grand Total: {qt_039.grand_total:,.2f})")

        await session.commit()

        print("\n── Copy PDFs to storage/imports ──")
        imp_dir = STORAGE_DIR / "imports"
        imp_dir.mkdir(parents=True, exist_ok=True)
        for pdf_file in DATA_DIR.glob("*.pdf"):
            dest = imp_dir / pdf_file.name
            if not dest.exists():
                shutil.copy2(pdf_file, dest)
                print(f"  Copied: {pdf_file.name}")
            else:
                print(f"  Already exists: {pdf_file.name}")

        print("\n✅ Seed complete!")
        print(f"   Brands: 4 | Categories: 4 | Customers: 5 | Projects: 5")
        print(f"   BOQs: 7 | Quotations: 3 (issued)")


if __name__ == "__main__":
    asyncio.run(main())
