"""
Import sales forecast/deal files from ../sale_data into dev DB.

Behavior:
- Maps each file to owner by username == filename stem (case-insensitive)
- Upserts customers/projects by name
- Creates deals for parsed rows
- Replaces monthly forecasts per created deal

Run from backend/:
  PYTHONPATH=. python scripts/import_sales_deals_from_excel.py

Optional:
  PYTHONPATH=. python scripts/import_sales_deals_from_excel.py --dry-run
"""
from __future__ import annotations

import argparse
import asyncio
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from sqlalchemy import delete, func, select

from app.database import AsyncSessionLocal
from app.models.customer import Customer
from app.models.deal import Deal
from app.models.deal_forecast import DealForecastMonthly
from app.models.project import Project
from app.models.user import User

BASE_DIR = Path(__file__).resolve().parents[1]
SALE_DATA_DIR_CANDIDATES = [
    BASE_DIR.parent / "sale_data",
    BASE_DIR / "sale_data",
]

MONTH_LABELS = {
    "jan": 1,
    "january": 1,
    "มกราคม": 1,
    "feb": 2,
    "february": 2,
    "กุมภาพันธ์": 2,
    "mar": 3,
    "march": 3,
    "มีนาคม": 3,
    "apr": 4,
    "april": 4,
    "เมษายน": 4,
    "may": 5,
    "พฤษภาคม": 5,
    "jun": 6,
    "june": 6,
    "มิถุนายน": 6,
    "jul": 7,
    "july": 7,
    "กรกฎาคม": 7,
    "aug": 8,
    "august": 8,
    "สิงหาคม": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "กันยายน": 9,
    "oct": 10,
    "october": 10,
    "ตุลาคม": 10,
    "nov": 11,
    "november": 11,
    "พฤศจิกายน": 11,
    "dec": 12,
    "december": 12,
    "ธันวาคม": 12,
}

STATUS_MAP: list[tuple[re.Pattern[str], tuple[str, str]]] = [
    (re.compile(r"award|won|ชนะ", re.IGNORECASE), ("won", "won")),
    (re.compile(r"lost|แพ้|cancel", re.IGNORECASE), ("lost", "lost")),
    (re.compile(r"hold|พัก", re.IGNORECASE), ("qualified", "on_hold")),
    (re.compile(r"negotia", re.IGNORECASE), ("negotiation", "open")),
    (re.compile(r"bidding|proposal|เสนอราคา|quote", re.IGNORECASE), ("proposal", "open")),
]


@dataclass
class ForecastItem:
    year: int
    month: int
    amount: Decimal


@dataclass
class ParsedDeal:
    customer_name: str
    project_name: str
    status_raw: str
    finish_raw: str
    remark: str
    monthly: list[ForecastItem] = field(default_factory=list)
    total_raw: Decimal = Decimal("0")
    win_pct_raw: Decimal | None = None
    net_total_raw: Decimal = Decimal("0")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split()).strip()


def to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    s = clean_text(value).replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def normalize_for_match(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip().lower()


def is_row_number(text: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.0+)?", (text or "").strip()))


def map_stage_and_status(raw_status: str) -> tuple[str, str]:
    text = raw_status or ""
    for pattern, mapped in STATUS_MAP:
        if pattern.search(text):
            return mapped
    return "lead", "open"


def parse_probability(win_pct_raw: Decimal | None, stage: str) -> int:
    if win_pct_raw is not None:
        pct = float(win_pct_raw)
        if pct <= 1:
            pct *= 100
        return int(max(0, min(100, round(pct))))
    defaults = {
        "lead": 10,
        "qualified": 30,
        "proposal": 50,
        "negotiation": 70,
        "won": 100,
        "lost": 0,
    }
    return defaults.get(stage, 10)


def parse_expected_close_date(finish_raw: str) -> date | None:
    text = (finish_raw or "").strip()
    if not text:
        return None

    m = re.search(r"Q([1-4])\s*/?\s*(20\d{2})", text, flags=re.IGNORECASE)
    if m:
        q = int(m.group(1))
        y = int(m.group(2))
        month = q * 3
        day = 31 if month in (3, 12) else 30
        return date(y, month, day)

    y = re.search(r"(20\d{2})", text)
    if y:
        return date(int(y.group(1)), 12, 31)

    return None


def map_months_openpyxl(sheet) -> tuple[dict[int, tuple[int, int]], int | None, int | None, int | None]:
    year_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    month_row = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    year_from_header = None
    month_cols: dict[int, tuple[int, int]] = {}
    total_col = None
    win_col = None
    net_col = None

    for idx, value in enumerate(year_row, start=1):
        text = clean_text(value)
        if not text:
            continue
        y = re.search(r"(20\d{2})", text)
        if y:
            year_from_header = int(y.group(1))

    for idx, value in enumerate(month_row, start=1):
        label = clean_text(value).lower()
        if not label:
            continue
        if label in MONTH_LABELS and year_from_header:
            month_cols[idx] = (year_from_header, MONTH_LABELS[label])
            continue
        if "total" in label and "%" not in label and "net" not in label:
            total_col = idx
        elif "win" in label or "%" in label:
            win_col = idx
        elif "net" in label:
            net_col = idx

    # 2027 annual bucket is often in U (col 21)
    if len(year_row) >= 21:
        y2027 = clean_text(year_row[20])
        if re.search(r"2027", y2027):
            month_cols[21] = (2027, 1)

    return month_cols, total_col, win_col, net_col


def extract_deals_from_openpyxl(ws) -> list[ParsedDeal]:
    month_cols, total_col, win_col, net_col = map_months_openpyxl(ws)
    out: list[ParsedDeal] = []
    current: ParsedDeal | None = None

    for r in range(5, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 30)]
        if not any(clean_text(v) for v in row):
            continue

        no_raw = clean_text(row[0])
        customer = clean_text(row[1])
        project = clean_text(row[2])
        status = clean_text(row[3])
        finish = clean_text(row[4])
        remark = clean_text(row[21]) if len(row) > 21 else ""

        is_new = is_row_number(no_raw) or (
            not no_raw and customer and project and bool(status)
        )

        if is_new:
            current = ParsedDeal(
                customer_name=customer or "Unknown Customer",
                project_name=project or f"Imported Project {r}",
                status_raw=status,
                finish_raw=finish,
                remark=remark,
            )
            out.append(current)
        elif current is None:
            continue

        assert current is not None

        if customer and not is_new and not customer.lower().startswith(("owner", "designer", "main con")):
            current.customer_name = customer
        if project and not is_new and not project.startswith("-"):
            current.project_name = project
        if status and not current.status_raw:
            current.status_raw = status
        if finish and not current.finish_raw:
            current.finish_raw = finish
        if remark:
            current.remark = "\n".join([x for x in [current.remark, remark] if x])

        for col, (year, month) in month_cols.items():
            if col - 1 >= len(row):
                continue
            amount = to_decimal(row[col - 1])
            if amount is None or amount == 0:
                continue
            current.monthly.append(ForecastItem(year=year, month=month, amount=amount * Decimal("1000000")))

        if total_col and total_col - 1 < len(row):
            total = to_decimal(row[total_col - 1])
            if total is not None:
                current.total_raw += total * Decimal("1000000")

        if win_col and win_col - 1 < len(row):
            win = to_decimal(row[win_col - 1])
            if win is not None and current.win_pct_raw is None:
                current.win_pct_raw = win

        if net_col and net_col - 1 < len(row):
            net = to_decimal(row[net_col - 1])
            if net is not None:
                current.net_total_raw += net * Decimal("1000000")

    # compact noisy rows
    cleaned: list[ParsedDeal] = []
    for d in out:
        if not d.customer_name and not d.project_name:
            continue
        if not d.monthly and d.total_raw == 0 and d.net_total_raw == 0:
            # Skip rows that are only notes without numeric forecast
            continue
        cleaned.append(d)

    return cleaned


def load_xlsx_deals(path: Path) -> list[ParsedDeal]:
    wb = openpyxl.load_workbook(path, data_only=True)
    preferred = ["Forcash_Update 23-03-26", "Forecast", "Sheet1", "Forcash_Old"]
    sheet = None
    for name in preferred:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        raise RuntimeError(f"No forecast-like sheet found in {path.name}")
    return extract_deals_from_openpyxl(sheet)


def load_xls_deals(path: Path) -> list[ParsedDeal]:
    book = xlrd.open_workbook(path.as_posix())
    target = None
    for s in book.sheets():
        n = s.name.lower()
        if "forecast" in n or "forcash" in n:
            target = s
            break
    if target is None:
        for s in book.sheets():
            probe = " ".join(
                clean_text(s.cell_value(r, c))
                for r in range(min(5, s.nrows))
                for c in range(min(5, s.ncols))
            ).lower()
            if "sales forecast" in probe:
                target = s
                break
    if target is None and book.nsheets > 0:
        target = book.sheet_by_index(0)
    if target is None:
        raise RuntimeError(f"No forecast-like sheet found in {path.name}")

    # Convert to a pseudo-openpyxl-like 2D matrix for reuse with lighter logic.
    # Header layout in xls is similar: row 3/4 for year + month labels.
    rows: list[list[Any]] = []
    for r in range(target.nrows):
        rows.append([target.cell_value(r, c) for c in range(min(target.ncols, 40))])

    year_row = rows[2] if len(rows) > 2 else []
    month_row = rows[3] if len(rows) > 3 else []

    year_from_header = None
    for v in year_row:
        m = re.search(r"(20\d{2})", clean_text(v))
        if m:
            year_from_header = int(m.group(1))
            break

    month_cols: dict[int, tuple[int, int]] = {}
    total_col = None
    win_col = None
    net_col = None
    for idx, v in enumerate(month_row):
        label = clean_text(v).lower()
        if not label:
            continue
        if label in MONTH_LABELS and year_from_header:
            month_cols[idx] = (year_from_header, MONTH_LABELS[label])
        elif "total" in label and "%" not in label and "net" not in label:
            total_col = idx
        elif "win" in label or "%" in label:
            win_col = idx
        elif "net" in label:
            net_col = idx

    out: list[ParsedDeal] = []
    current: ParsedDeal | None = None
    current_customer = ""

    for r in range(4, len(rows)):
        row = rows[r]
        if not any(clean_text(v) for v in row):
            continue

        no_raw = clean_text(row[0]) if len(row) > 0 else ""
        customer = clean_text(row[1]) if len(row) > 1 else ""
        project = clean_text(row[2]) if len(row) > 2 else ""
        status = clean_text(row[3]) if len(row) > 3 else ""
        finish = clean_text(row[4]) if len(row) > 4 else ""
        remark = clean_text(row[21]) if len(row) > 21 else ""

        if customer:
            current_customer = customer

        is_new = is_row_number(no_raw) or (
            not no_raw and customer and project and bool(status)
        )

        if is_new:
            base_customer = customer or current_customer or "Unknown Customer"
            current = ParsedDeal(
                customer_name=base_customer,
                project_name=project or "",
                status_raw=status,
                finish_raw=finish,
                remark=remark,
            )
            if current.project_name or current.status_raw:
                out.append(current)
            else:
                # Header-like holder row (customer only). Keep as current context but do not emit deal yet.
                current = None
                continue
        elif current is None:
            continue

        assert current is not None

        if customer and not current.customer_name:
            current.customer_name = customer
        if project and (not current.project_name or current.project_name.startswith("-")):
            current.project_name = project
        if status and not current.status_raw:
            current.status_raw = status
        if finish and not current.finish_raw:
            current.finish_raw = finish
        if remark:
            current.remark = "\n".join([x for x in [current.remark, remark] if x])

        for col, (year, month) in month_cols.items():
            if col >= len(row):
                continue
            amount = to_decimal(row[col])
            if amount is None or amount == 0:
                continue
            current.monthly.append(ForecastItem(year=year, month=month, amount=amount * Decimal("1000000")))

        if total_col is not None and total_col < len(row):
            total = to_decimal(row[total_col])
            if total is not None:
                current.total_raw += total * Decimal("1000000")

        if win_col is not None and win_col < len(row) and current.win_pct_raw is None:
            win = to_decimal(row[win_col])
            if win is not None:
                current.win_pct_raw = win

        if net_col is not None and net_col < len(row):
            net = to_decimal(row[net_col])
            if net is not None:
                current.net_total_raw += net * Decimal("1000000")

    cleaned: list[ParsedDeal] = []
    for d in out:
        if not d.project_name:
            continue
        if not d.monthly and d.total_raw == 0 and d.net_total_raw == 0:
            continue
        cleaned.append(d)
    return cleaned


async def upsert_customer(session, name: str) -> Customer:
    q = await session.execute(
        select(Customer).where(func.lower(Customer.name) == normalize_for_match(name))
    )
    obj = q.scalar_one_or_none()
    if obj:
        return obj
    obj = Customer(name=name)
    session.add(obj)
    await session.flush()
    return obj


async def upsert_project(session, customer_id: int, name: str) -> Project:
    q = await session.execute(
        select(Project).where(func.lower(Project.name) == normalize_for_match(name))
    )
    obj = q.scalar_one_or_none()
    if obj:
        if obj.customer_id != customer_id:
            obj.customer_id = customer_id
        return obj
    obj = Project(customer_id=customer_id, name=name, status="active")
    session.add(obj)
    await session.flush()
    return obj


def aggregate_monthly(items: list[ForecastItem]) -> list[ForecastItem]:
    agg: dict[tuple[int, int], Decimal] = {}
    for i in items:
        key = (i.year, i.month)
        agg[key] = agg.get(key, Decimal("0")) + i.amount
    out = [ForecastItem(year=y, month=m, amount=a) for (y, m), a in sorted(agg.items()) if a != 0]
    return out


async def import_file(path: Path, dry_run: bool = False) -> dict[str, int]:
    owner_name = path.stem.lower()

    async with AsyncSessionLocal() as session:
        uq = await session.execute(select(User).where(func.lower(User.username) == owner_name))
        owner = uq.scalar_one_or_none()
        if not owner:
            raise RuntimeError(f"Owner user not found for file {path.name}: username={owner_name}")

        if path.suffix.lower() == ".xlsx":
            parsed = load_xlsx_deals(path)
        elif path.suffix.lower() == ".xls":
            parsed = load_xls_deals(path)
        else:
            return {"deals": 0, "forecasts": 0}

        created_deals = 0
        created_forecasts = 0

        for row in parsed:
            customer_name = row.customer_name[:200] if row.customer_name else "Unknown Customer"
            project_name = row.project_name[:255] if row.project_name else "Imported Project"

            customer = await upsert_customer(session, customer_name)
            project = await upsert_project(session, customer.id, project_name)

            stage, status = map_stage_and_status(row.status_raw)
            probability = parse_probability(row.win_pct_raw, stage)

            monthly = aggregate_monthly(row.monthly)

            expected_value = Decimal("0")
            if monthly:
                expected_value = sum((m.amount for m in monthly), Decimal("0"))
            elif row.net_total_raw:
                expected_value = row.net_total_raw
            elif row.total_raw:
                expected_value = row.total_raw

            note_parts = [
                f"import_source={path.name}",
                f"status_raw={row.status_raw}" if row.status_raw else "",
                f"finish_raw={row.finish_raw}" if row.finish_raw else "",
                row.remark,
            ]
            description = "\n".join([p for p in note_parts if p]).strip() or None

            deal = Deal(
                title=project_name,
                customer_id=customer.id,
                project_id=project.id,
                owner_id=owner.id,
                deal_cycle_stage=stage,
                status=status,
                expected_value=expected_value,
                probability_pct=probability,
                expected_close_date=parse_expected_close_date(row.finish_raw),
                source="excel_import",
                description=description,
            )
            session.add(deal)
            await session.flush()
            created_deals += 1

            if monthly:
                for m in monthly:
                    net = (m.amount * Decimal(probability) / Decimal("100")).quantize(Decimal("0.01"))
                    session.add(
                        DealForecastMonthly(
                            deal_id=deal.id,
                            forecast_year=m.year,
                            forecast_month=m.month,
                            amount=m.amount,
                            win_pct=Decimal(probability),
                            net_amount=net,
                            note=f"import_source={path.name}",
                        )
                    )
                    created_forecasts += 1

        if dry_run:
            await session.rollback()
        else:
            await session.commit()

    return {"deals": created_deals, "forecasts": created_forecasts}


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only, then rollback")
    args = parser.parse_args()

    sale_data_dir = next((p for p in SALE_DATA_DIR_CANDIDATES if p.exists()), None)
    if sale_data_dir is None:
        raise RuntimeError(f"No sale_data directory found in: {SALE_DATA_DIR_CANDIDATES}")

    files = sorted([*sale_data_dir.glob("*.xlsx"), *sale_data_dir.glob("*.xls")])
    if not files:
        raise RuntimeError(f"No files found in {sale_data_dir}")

    total_deals = 0
    total_forecasts = 0

    print(f"Import files: {[f.name for f in files]}")
    for f in files:
        summary = await import_file(f, dry_run=args.dry_run)
        total_deals += summary["deals"]
        total_forecasts += summary["forecasts"]
        print(f"{f.name}: deals={summary['deals']} monthly_forecasts={summary['forecasts']}")

    mode = "DRY RUN" if args.dry_run else "COMMIT"
    print(f"Done ({mode}) total_deals={total_deals} total_monthly_forecasts={total_forecasts}")


if __name__ == "__main__":
    asyncio.run(main())
